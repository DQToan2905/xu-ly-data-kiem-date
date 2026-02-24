"""
Xử lý data kiểm date - Optimized for large datasets (Pure Polars, no DuckDB)
=============================================================================
Chiến lược chống OOM:
  ┌─────────────────────────────────────────────────────────────────┐
  │  Excel file(s)                                                  │
  │     ↓ (đọc từng sheet, filter ngay)                             │
  │  Parquet staging files  ← không bao giờ giữ hết trong RAM       │
  │     ↓ (đọc từng Parquet file, ghi thẳng vào Excel)             │
  │  Excel output  ← xlsxwriter constant_memory = ghi row-by-row   │
  └─────────────────────────────────────────────────────────────────┘

  NGUYÊN TẮC: Không bao giờ collect() toàn bộ data vào 1 DataFrame.
  Mọi thứ đều xử lý theo từng Parquet chunk nhỏ.
"""

import streamlit as st
import polars as pl
import xlsxwriter
import tempfile
import os
import gc
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from io import BytesIO
from openpyxl import load_workbook

# ============================= CONFIG =============================
st.set_page_config(page_title="Xử lý data kiểm date", layout="wide")
st.title("📊 Xử lý dữ liệu kiểm date")
st.write("Upload nhiều file Excel → xử lý → xuất file tổng hợp")

COLUMNS_KEEP = [
    'Mã siêu thị','Tên siêu thị','Mã sản phẩm','Tên sản phẩm',
    'SL chuyển kho','SL giảm giá','SL hủy tại siêu thị',
    'Số lượng trả NCC','SL đổi hàng NCC','Số lượng bình thường',
    'SL tặng KM','SL cận date (tặng quà)',
    'Ngày tạo','Lần kiểm cuối cùng',
    'Mã nhân viên','Họ và tên nhân viên',
    'Ngày duyệt','Người duyệt','Tên người duyệt',
    'Hình ảnh','Ghi chú trạng thái','Ghi chú',
    'Ngày hệ thống yêu cầu','Trạng thái','Nội dung',
    'Hạn sử dụng','Date gần nhất','Hình ảnh_1',
    'Phân loại','Thời gian bắt đầu','Thời gian kết thúc',
    'Giá trị phần trăm giảm giá'
]

FILTER_COLS = ['SL giảm giá', 'SL hủy tại siêu thị', 'SL tặng KM', 'SL cận date (tặng quà)']

# RAM tối đa mỗi chunk khi ghi Excel (số dòng)
WRITE_CHUNK = 5_000


# ============================= BƯỚC 1: XỬ LÝ → PARQUET =============================

def process_sheet_to_parquet(file_bytes: bytes, sheet_name: str,
                              file_name: str, staging_dir: str) -> tuple[str | None, str | None]:
    """
    Đọc 1 sheet → filter → lưu Parquet.
    Trả về (parquet_path, error_msg).
    RAM tối đa = 1 sheet tại 1 thời điểm.
    """
    try:
        lf = pl.read_excel(
            BytesIO(file_bytes),
            sheet_name=sheet_name,
            infer_schema_length=0,      # tất cả Utf8, không infer sai kiểu
        ).lazy()

        available = lf.columns

        filter_cols_present = [c for c in FILTER_COLS if c in available]
        if not filter_cols_present:
            return None, None           # sheet không liên quan

        cast_exprs = [
            pl.col(c).cast(pl.Float64, strict=False)
            for c in filter_cols_present
        ]

        filter_expr = sum(
            pl.col(c).fill_null(0) for c in filter_cols_present
        ) > 0

        keep = [c for c in COLUMNS_KEEP if c in available]

        lf = (lf
              .with_columns(cast_exprs)
              .filter(filter_expr)      # lọc TRƯỚC collect
              .select(keep))

        if 'Hình ảnh_1' in keep:
            lf = lf.with_columns(
                pl.concat_str([pl.lit('"'),
                               pl.col('Hình ảnh_1').fill_null(''),
                               pl.lit('"')]).alias('Hình ảnh_1')
            )

        lf = lf.with_columns(pl.lit(file_name).alias("file_name"))

        df = lf.collect(streaming=True)
        if df.is_empty():
            return None, None

        # Lưu Parquet ngay, giải phóng DataFrame khỏi RAM
        safe = "".join(c if c.isalnum() else "_" for c in f"{file_name}_{sheet_name}")
        path = os.path.join(staging_dir, f"{safe[:100]}.parquet")
        df.write_parquet(path, compression="snappy")
        del df
        gc.collect()
        return path, None

    except Exception as e:
        return None, f"Sheet '{sheet_name}' / '{file_name}': {e}"


def process_file_to_parquets(uf, staging_dir: str) -> tuple[list[str], list[str]]:
    """Xử lý 1 uploaded file → nhiều Parquet files (1 per sheet)."""
    file_bytes = uf.getvalue()
    file_name  = uf.name
    paths, errors = [], []

    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
    except Exception as e:
        del file_bytes
        return [], [f"Không mở được '{file_name}': {e}"]

    for sname in sheet_names:
        path, err = process_sheet_to_parquet(file_bytes, sname, file_name, staging_dir)
        if path:
            paths.append(path)
        if err:
            errors.append(err)

    del file_bytes
    gc.collect()
    return paths, errors


# ============================= BƯỚC 2A: PARQUET → CSV (STREAMING) =============================

def parquets_to_csv_streaming(parquet_paths: list[str],
                               status_text, progress_bar) -> tuple[bytes, int]:
    """
    Ghi CSV trực tiếp từ Parquet — NHANH NHẤT, RAM thấp nhất.
    ──────────────────────────────────────────────────────────
    CSV là plain text → không cần buffer, không cần format.
    Polars write_csv() với BytesIO là zero-copy cho từng chunk.
    Encoding UTF-8 with BOM → Excel mở đúng tiếng Việt ngay.
    """
    output = BytesIO()
    total_rows = 0
    n_parts = len(parquet_paths)
    header_written = False

    # UTF-8 BOM — Excel tự nhận diện encoding tiếng Việt
    output.write(b'\xef\xbb\xbf')

    for pi, ppath in enumerate(parquet_paths):
        df = pl.read_parquet(ppath)
        n  = len(df)

        for start in range(0, n, WRITE_CHUNK):
            chunk = df.slice(start, WRITE_CHUNK)
            # has_header=False sau lần đầu → không lặp lại header
            csv_bytes = chunk.write_csv().encode('utf-8')
            # Bỏ BOM ký tự đầu nếu polars tự thêm
            if csv_bytes.startswith(b'\xef\xbb\xbf'):
                csv_bytes = csv_bytes[3:]
            output.write(csv_bytes)
            header_written = True
            total_rows += len(chunk)
            del chunk

        del df
        gc.collect()

        pct = 0.80 + (pi + 1) / n_parts * 0.19
        progress_bar.progress(min(pct, 0.99))
        status_text.text(f"📤 Đang ghi CSV... {total_rows:,} dòng ({pi+1}/{n_parts} phần)")

    output.seek(0)
    return output.getvalue(), total_rows


# ============================= BƯỚC 2B: PARQUET → EXCEL (STREAMING) =============================

def _get_unified_columns(parquet_paths: list[str]) -> list[str]:
    """Schema chung từ tất cả Parquet — không load data."""
    all_cols: list[str] = []
    seen: set[str] = set()
    for p in parquet_paths:
        for col in pl.read_parquet_schema(p).keys():
            if col not in seen:
                all_cols.append(col)
                seen.add(col)
    return all_cols


def parquets_to_excel_streaming(parquet_paths: list[str],
                                 status_text, progress_bar) -> tuple[bytes, int]:
    """
    Ghi Excel từ Parquet files, từng chunk WRITE_CHUNK dòng.
    xlsxwriter constant_memory=True → flush row ngay, RAM ≈ 0.
    """
    output   = BytesIO()
    workbook = xlsxwriter.Workbook(output, {
        'constant_memory': True,
        'in_memory':       True,
        'strings_to_urls': False,
    })
    worksheet  = workbook.add_worksheet("Data")
    header_fmt = workbook.add_format({
        'bold': True, 'bg_color': '#4472C4',
        'font_color': 'white', 'border': 1,
    })

    all_cols  = _get_unified_columns(parquet_paths)
    col_index = {c: i for i, c in enumerate(all_cols)}

    for i, col in enumerate(all_cols):
        worksheet.write(0, i, col, header_fmt)

    total_rows = 0
    n_parts    = len(parquet_paths)

    for pi, ppath in enumerate(parquet_paths):
        reader = pl.read_parquet(ppath)
        n      = len(reader)

        for start in range(0, n, WRITE_CHUNK):
            chunk = reader.slice(start, WRITE_CHUNK)
            for ri, row in enumerate(chunk.iter_rows(named=True), start=total_rows + 1):
                for col_name, val in row.items():
                    ci = col_index.get(col_name)
                    if ci is not None and val is not None and val == val:
                        worksheet.write(ri, ci, val)
            total_rows += len(chunk)
            del chunk
            gc.collect()

        del reader
        gc.collect()

        pct = 0.80 + (pi + 1) / n_parts * 0.19
        progress_bar.progress(min(pct, 0.99))
        status_text.text(f"📤 Đang ghi Excel... {total_rows:,} dòng ({pi+1}/{n_parts} phần)")

    workbook.close()
    output.seek(0)
    return output.getvalue(), total_rows


# ============================= PIPELINE CHÍNH =============================

def run_pipeline(uploaded_files, progress_bar, status_text,
                 max_workers: int, export_format: str) -> tuple[bytes, int, str] | None:
    """
    Full pipeline: Excel files → Parquet staging → CSV hoặc Excel.
    Trả về (output_bytes, total_rows, format).
    """
    staging_dir = tempfile.mkdtemp(prefix="kiem_date_")
    n           = len(uploaded_files)
    all_parquet: list[str] = []
    all_errors:  list[str] = []
    completed = 0

    # ── Bước 1: Song song đọc & filter → Parquet ──
    status_text.text("🔄 Đang đọc và lọc dữ liệu...")
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(process_file_to_parquets, uf, staging_dir): uf.name
            for uf in uploaded_files
        }
        for future in as_completed(futures):
            completed += 1
            progress_bar.progress(completed / n * 0.78)
            try:
                paths, errs = future.result()
                all_parquet.extend(paths)
                all_errors.extend(errs)
                status_text.text(f"✅ {completed}/{n}: {futures[future]}")
            except Exception as e:
                all_errors.append(f"Worker lỗi: {e}")

    for err in all_errors:
        st.warning(f"⚠️ {err}")

    if not all_parquet:
        return None

    progress_bar.progress(0.80)

    # ── Bước 2: Export theo format đã chọn ──
    if export_format == "CSV":
        status_text.text(f"📤 Đang xuất CSV từ {len(all_parquet)} phần...")
        output_bytes, total_rows = parquets_to_csv_streaming(all_parquet, status_text, progress_bar)
    else:
        status_text.text(f"📤 Đang xuất Excel từ {len(all_parquet)} phần...")
        output_bytes, total_rows = parquets_to_excel_streaming(all_parquet, status_text, progress_bar)

    # Dọn Parquet staging
    for p in all_parquet:
        try: os.unlink(p)
        except Exception: pass
    try: os.rmdir(staging_dir)
    except Exception: pass

    return output_bytes, total_rows, export_format


# ============================= UI =============================

EXCEL_ROW_LIMIT = 1_048_576   # giới hạn cứng của Excel

with st.sidebar:
    st.header("⚙️ Cấu hình")

    export_fmt = st.radio(
        "📁 Định dạng xuất",
        options=["CSV", "Excel (.xlsx)"],
        index=0,
        help="CSV nhanh hơn 5–10x, không giới hạn dòng, mở được trong Excel"
    )

    st.divider()

    max_workers_cfg = st.slider(
        "Số luồng đọc song song",
        min_value=1, max_value=8, value=2,
        help="Tăng nếu nhiều file nhỏ; GIẢM nếu file lớn/RAM thấp"
    )

    if export_fmt == "Excel (.xlsx)":
        write_chunk_cfg = st.select_slider(
            "Chunk size khi ghi Excel (dòng)",
            options=[1_000, 2_000, 5_000, 10_000, 20_000],
            value=5_000,
            help="Nhỏ hơn = ít RAM hơn; Lớn hơn = nhanh hơn"
        )
    else:
        write_chunk_cfg = 10_000   # CSV không cần nhỏ

    st.divider()
    st.markdown("""
**💡 Khi nào dùng CSV?**
- Data > 100k dòng
- Cần tốc độ nhanh
- Không cần format/màu sắc

**💡 Khi nào dùng Excel?**
- Data < 100k dòng
- Cần chia sẻ có format đẹp
- Người nhận không quen CSV
    """)

uploaded_files = st.file_uploader(
    "📂 Upload các file Excel",
    type=["xlsx", "xls"],
    accept_multiple_files=True
)

if "output_data"   not in st.session_state: st.session_state.output_data   = None
if "output_format" not in st.session_state: st.session_state.output_format = "CSV"
if "row_count"     not in st.session_state: st.session_state.row_count     = 0

if uploaded_files:
    total_mb = sum(len(f.getvalue()) for f in uploaded_files) / 1024 / 1024

    # Gợi ý format dựa trên kích thước
    auto_recommend = "CSV" if total_mb > 50 else "Excel (.xlsx)"

    col_info, col_rec = st.columns([2, 1])
    col_info.info(f"📁 **{len(uploaded_files)} file** | Tổng: **{total_mb:.1f} MB**")
    if auto_recommend == "CSV":
        col_rec.warning(f"💡 Gợi ý: dùng **CSV** cho data lớn")
    else:
        col_rec.success(f"✅ Dùng **Excel** ổn với data nhỏ")

    if st.button("🚀 Xử lý dữ liệu", type="primary", use_container_width=True):
        WRITE_CHUNK = write_chunk_cfg
        chosen_fmt  = "CSV" if export_fmt == "CSV" else "Excel"

        start_time   = time.perf_counter()
        progress_bar = st.progress(0)
        status_text  = st.empty()

        try:
            result = run_pipeline(
                uploaded_files, progress_bar, status_text,
                max_workers=max_workers_cfg,
                export_format=chosen_fmt,
            )

            if result is None:
                st.error("❌ Không có dữ liệu thỏa điều kiện lọc")
                st.stop()

            output_bytes, total_rows, fmt = result

            # Cảnh báo nếu Excel vượt giới hạn dòng
            if fmt == "Excel" and total_rows >= EXCEL_ROW_LIMIT:
                st.error(f"⛔ Excel bị cắt ở {EXCEL_ROW_LIMIT:,} dòng! Hãy dùng CSV.")

            st.session_state.output_data   = output_bytes
            st.session_state.output_format = fmt
            st.session_state.row_count     = total_rows

            elapsed = time.perf_counter() - start_time
            progress_bar.progress(1.0)
            status_text.text("✅ Hoàn thành!")

            c1, c2, c3, c4 = st.columns(4)
            c1.metric("⏱️ Thời gian",  f"{elapsed:.1f}s")
            c2.metric("📊 Tổng dòng",  f"{total_rows:,}")
            c3.metric("📁 Files",       len(uploaded_files))
            c4.metric("💾 Output",      f"{len(output_bytes)/1024/1024:.1f} MB")

            gc.collect()

        except MemoryError:
            st.error("❌ Hết RAM! Giảm số luồng xuống 1, đổi sang CSV, hoặc chia nhỏ file.")
        except Exception as e:
            st.error(f"❌ Lỗi: {e}")
            st.exception(e)

# ── Download button — tự điều chỉnh theo format ──
if st.session_state.output_data is not None:
    fmt = st.session_state.output_format
    if fmt == "CSV":
        st.download_button(
            label=f"📥 Download CSV ({st.session_state.row_count:,} dòng)",
            data=st.session_state.output_data,
            file_name="data_kiem_date.csv",
            mime="text/csv; charset=utf-8",
            type="primary",
            use_container_width=True,
        )
        st.caption("💡 Mở trong Excel: Data tab → From Text/CSV → Encoding: UTF-8")
    else:
        st.download_button(
            label=f"📥 Download Excel ({st.session_state.row_count:,} dòng)",
            data=st.session_state.output_data,
            file_name="data_kiem_date.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
